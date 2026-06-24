import math
import sqlite3
import tempfile
import unittest
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook

import database as db


def create_maroc_workbook(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tableau COE"
    sheet.append([
        "N°",
        "SEXE",
        "NOM ET PRENOM (S)",
        "DATE & LIEU DE NAISSANCE",
        "DIPLOME / FILIERE / ANNEE",
        "MOYENNE / MENTION",
        "OBSERVATION",
        "AVIS CNaBAU",
    ])
    sheet.append(["NIVEAU : LICENCE"])
    sheet.append(["Filière : Sciences de l'Ingénieur (02 places)"])
    sheet.append([
        1,
        "F",
        "CANDIDATE TEST",
        "01/01/2008 à Cotonou",
        "BAC/C/2025",
        "16,25 Très-bien",
        "RAS",
        None,
    ])
    sheet.append(["NIVEAU : SPECIALITE MEDICALE"])
    sheet.append(["Filière : Filière : Pneumologie (01 place)"])
    sheet.append([
        2,
        "M",
        "CANDIDAT TEST",
        "01/01/1999 à Porto-Novo",
        "DOCTORAT EN MEDECINE/2025",
        None,
        None,
        None,
    ])
    workbook.save(path)


def create_professional_workbook(path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tableau COE"
    sheet.append([
        "N°",
        "SEXE",
        "NOM ET PRENOM (S)",
        "DATE & LIEU DE NAISSANCE",
        "DIPLOME / FILIERE / ANNEE",
        "MOYENNE / MENTION",
        "OBSERVATION",
        "AVIS CNaBAU",
    ])
    sheet.append(["NIVEAU : BAC + 2 ans"])
    sheet.append(["Filière : Génie Electrique (02 places)"])
    sheet.append([
        1,
        "M",
        "PREMIER DOSSIER",
        "01/01/2006 à Cotonou",
        "BAC D 2025",
        "13,00 Assez bien",
        "RAS",
        None,
    ])
    sheet.append([
        2,
        "F",
        "SECOND DOSSIER",
        "02/02/2006 à Porto-Novo",
        "BAC C 2025",
        "14,00 Bien",
        "RAS",
        "BAC ETRANGER",
    ])
    sheet.append([
        1,
        "F",
        "DOUBLON À IGNORER",
        "03/03/2006 à Parakou",
        "BAC D 2025",
        "12,00 Assez bien",
        "RAS",
        None,
    ])
    workbook.save(path)


class MarocImportTest(unittest.TestCase):
    def test_imports_candidates_and_quotas(self):
        original_db_path = db.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp:
                db.DB_PATH = str(Path(tmp) / "session.db")
                workbook_path = Path(tmp) / "maroc.xlsx"
                create_maroc_workbook(workbook_path)
                db.init_db()

                count = db.load_excel_to_db(str(workbook_path))
                candidates = db.get_all_candidatures()
                quotas = db.get_quotas()

                self.assertEqual(count, 2)
                self.assertEqual(len(candidates), 2)
                self.assertEqual(len(quotas), 2)
                self.assertEqual(sum(quotas.values()), 3)
                self.assertEqual(candidates["numero"].nunique(), 2)
                self.assertEqual(
                    set(candidates["niveau_etudes"]),
                    {"Licence", "Spécialité médicale"},
                )
                self.assertTrue(
                    all(
                        (row.niveau_etudes, row.filiere) in quotas
                        for row in candidates.itertuples()
                    )
                )
                first = candidates.sort_values("numero").iloc[0]
                self.assertEqual(first["id_demande"], "MAR-0001/26")
                self.assertEqual(first["name"], "CANDIDATE TEST")
        finally:
            db.DB_PATH = original_db_path

    def test_imports_professional_file_and_keeps_first_duplicate_number(self):
        original_db_path = db.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp:
                db.DB_PATH = str(Path(tmp) / "session.db")
                workbook_path = Path(tmp) / "maroc_fp.xlsx"
                create_professional_workbook(workbook_path)
                db.init_db()

                count = db.load_excel_to_db(str(workbook_path))
                candidates = db.get_all_candidatures().sort_values("numero")
                quotas = db.get_quotas()

                self.assertEqual(count, 2)
                self.assertEqual(len(candidates), 2)
                self.assertEqual(set(candidates["niveau_etudes"]), {"Bac + 2 ans"})
                self.assertEqual(candidates.iloc[0]["name"], "PREMIER DOSSIER")
                self.assertEqual(candidates.iloc[1]["avis"], "En attente")
                self.assertIn("BAC ETRANGER", candidates.iloc[1]["observation"])
                self.assertEqual(quotas, {("Bac + 2 ans", "Génie Electrique"): 2})
        finally:
            db.DB_PATH = original_db_path

    def test_assigns_reorders_and_exports_suppleant_ranks(self):
        original_db_path = db.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp:
                db.DB_PATH = str(Path(tmp) / "session.db")
                workbook_path = Path(tmp) / "maroc_fp.xlsx"
                create_professional_workbook(workbook_path)
                db.init_db()
                db.load_excel_to_db(str(workbook_path))

                db.update_avis("MAR-0001/26", "Suppléant")
                db.update_avis("MAR-0002/26", "Suppléant")
                candidates = db.get_all_candidatures().sort_values("numero")
                self.assertEqual(list(candidates["rang_suppleant"]), [1, 2])

                db.update_avis("MAR-0001/26", "Favorable")
                candidates = db.get_all_candidatures().sort_values("numero")
                self.assertTrue(math.isnan(candidates.iloc[0]["rang_suppleant"]))
                self.assertEqual(candidates.iloc[1]["rang_suppleant"], 1)

                db.update_avis("MAR-0001/26", "Suppléant")
                candidates = db.get_all_candidatures().sort_values("numero")
                self.assertEqual(list(candidates["rang_suppleant"]), [2, 1])

                excel_path = Path(tmp) / "decisions.xlsx"
                word_path = Path(tmp) / "decisions.docx"
                db.export_avis_to_xlsx(str(excel_path))
                db.export_to_docx(str(word_path))

                exported = load_workbook(excel_path, data_only=True)
                sheet = exported["Suppléants"]
                rank_values = [
                    cell.value
                    for row in sheet.iter_rows(min_row=2)
                    for cell in [row[1]]
                    if cell.value
                ]
                self.assertEqual(rank_values, ["1er suppléant", "2e suppléant"])

                document = Document(word_path)
                text = "\n".join(
                    cell.text
                    for table in document.tables
                    for row in table.rows
                    for cell in row.cells
                )
                self.assertIn("1er suppléant", text)
                self.assertIn("2e suppléant", text)
        finally:
            db.DB_PATH = original_db_path

    def test_migrates_existing_database_and_assigns_missing_ranks(self):
        original_db_path = db.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp:
                db.DB_PATH = str(Path(tmp) / "legacy.db")
                connection = sqlite3.connect(db.DB_PATH)
                connection.execute("""
                    CREATE TABLE candidatures (
                        id_demande TEXT PRIMARY KEY,
                        id_russe TEXT,
                        numero INTEGER,
                        sexe TEXT,
                        name TEXT NOT NULL,
                        date_lieu_naissance TEXT,
                        diplome_filiere_annee TEXT,
                        moyenne TEXT,
                        observation TEXT,
                        filiere TEXT NOT NULL,
                        niveau_etudes TEXT NOT NULL,
                        avis TEXT DEFAULT 'En attente'
                    )
                """)
                connection.executemany(
                    """INSERT INTO candidatures
                       (id_demande, numero, name, filiere, niveau_etudes, avis)
                       VALUES (?, ?, ?, ?, ?, 'Suppléant')""",
                    [
                        ("MAR-0004/26", 4, "QUATRIEME", "Génie Electrique", "Bac + 2 ans"),
                        ("MAR-0003/26", 3, "TROISIEME", "Génie Electrique", "Bac + 2 ans"),
                    ],
                )
                connection.commit()
                connection.close()

                db.init_db()
                candidates = db.get_all_candidatures().sort_values("numero")
                self.assertEqual(list(candidates["rang_suppleant"]), [1, 2])
        finally:
            db.DB_PATH = original_db_path


if __name__ == "__main__":
    unittest.main()
